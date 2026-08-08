#!/usr/bin/env python3
"""Audit (and optionally repair) the design workbook against the game source.

    python tools/design-audit.py                 # report drift, exit 1 if any
    python tools/design-audit.py --fix           # also correct the wrong cells
    python tools/design-audit.py --sheet Enemies # narrow to one sheet
    python tools/design-audit.py --workbook path/to/Game1.xlsx

The numbers in Game1.xlsx are a snapshot of tables that live in the JS. They go
stale silently: a region gets inserted and every later Rgn # is wrong, a potion
is renamed, an enemy loses its `ranged` flag. This walks the derivable sheets
cell by cell against tools/design-export.js (which runs the real game code) and
names every disagreement.

WHAT IT CAN AND CANNOT DO
  * MISMATCH  — a cell whose value disagrees with the code. `--fix` rewrites it.
  * MISSING   — a row the code has and the sheet does not. The full row is
                printed, but never inserted: placement (which banded section,
                what sort order) and the sheet's hand-authored columns like
                Foliage "Type" need a human call. See the design-workbook skill
                for the style-preserving insert recipe.
  * EXTRA     — a sheet row matching nothing in the code. Never auto-deleted.
  * CONSTANT  — a prose sheet quotes a source constant, and the sentence no
                longer matches. Anchored on the wording, not just the number, so
                it does not pass on a coincidental digit elsewhere on the sheet.
                It only covers constants listed in PROSE_CONSTANTS; prose that
                went stale around an unchanged number still needs a reader.

Three sheets are prose (Treasure & Chests, Quests, NPCs & Services) and get the
constant scan only — their rules live in function bodies, not tables, so read the
source listed in the skill when you touch them.

Requires: node on PATH, openpyxl.
"""

import argparse
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit('design-audit: openpyxl is not installed - pip install openpyxl')

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = ROOT.parent / 'Game1.xlsx'
NONE = '—'          # em dash: the workbook's "not applicable" marker


# ─── Sheet specs ──────────────────────────────────────────────────────────────
# Each entry maps a table in the workbook to a list in the export. `columns` is
# {sheet column header: export field}; a header not listed here is left alone,
# which is how hand-authored columns (Foliage "Type", every Notes column) opt out
# of the audit. `key` is the column whose value identifies a row on both sides.
# `section` narrows a table to the rows under a banner, for sheets that stack two
# tables under one header row.

class Table:
    def __init__(self, sheet, source, key, key_field, columns,
                 section=None, header_contains=None):
        self.sheet = sheet
        self.source = source                # key into the export JSON
        self.key = key                      # sheet column header used as identity
        self.key_field = key_field          # matching export field
        self.columns = columns
        self.section = section              # banner text the table sits under
        self.header_contains = header_contains or key


TABLES = [
    Table('Enemies', 'enemies', 'Enemy', 'name', {
        'Region': 'region', 'Rgn #': 'regionNum', 'Class': 'class',
        'HP': 'hp', 'Damage': 'damage', 'Dmg Element': 'element',
        'Vulnerability': 'vulnerability', 'Swims': 'swims',
        'Trophy Drop': 'trophy', 'Drop %': 'trophyPct',
        'Trophy Sell (rubies)': 'trophySell', 'Other Drops': 'otherDrops',
        'XP Awarded': 'xp', 'Speed (ms, lower=faster)': 'speed',
        'Size': 'size', 'CR': 'cr',
    }),
    # "Type" (Bloom / Fern / Mineral / …) is a hand-authored taxonomy with no
    # counterpart in the code, so it is deliberately absent below.
    Table('Foliage', 'foliage', 'Foliage', 'foliage', {
        'Region': 'region', 'Rgn #': 'regionNum', 'Foliage': 'foliage',
        'Reverts To': 'revertsTo', 'Drop Item': 'drop', 'Drop %': 'dropPct',
        'Sell (rubies)': 'sell', 'Sold Where': 'soldWhere',
        'Herbalist Use': 'herbalistUse', 'Region Potion': 'potionDice',
    }),
    Table('Regions', 'regions', 'Region', 'region', {
        'Rgn #': 'regionNum', 'Region': 'region', 'Element': 'element',
        'Enemy Tier': 'enemyTier', 'Boss': 'boss', 'Village': 'village',
        'Ore Yielded': 'ore',
    }),
    Table('Ores & Minerals', 'ores', 'Ore', 'ore', {
        'Tier': 'tier', 'Ore': 'ore', 'Icon': 'icon',
        'Sell Value (rubies)': 'sell', 'Regions (index)': 'regions',
        'Blacksmith Armor Bonus': 'armorBonus',
    }),
    Table('Progression', 'progression', 'Level', 'level', {
        'Level': 'level', 'XP to Next Level': 'xpToNext',
        'Cumulative XP': 'cumulative', 'Max HP at Level': 'maxHp',
    }),
    Table('Sword Progression', 'swordForge', 'Element', 'element', {
        'Element': 'element', 'Rgn #': 'regionNum', 'Home Region': 'homeRegion',
        'Forge Cost (rubies)': 'forgeCost',
    }, header_contains='Forge Cost (rubies)'),
    Table('Sword Progression', 'ladder', 'To Level', 'level', {
        'To Level': 'level', 'Ore Tier Used': 'ore', 'Ore Cost': 'oreCost',
        'Ruby Cost': 'rubyCost', 'Elemental Damage': 'damage',
    }, header_contains='Elemental Damage'),
    Table('Armor Progression', 'armorForge', 'Element', 'element', {
        'Element': 'element', 'Rgn #': 'regionNum', 'Home Region': 'homeRegion',
        'Forge Cost (rubies)': 'forgeCost',
    }, header_contains='Forge Cost (rubies)'),
    Table('Armor Progression', 'ladder', 'To Level', 'level', {
        'To Level': 'level', 'Ore Tier Used': 'ore', 'Ore Cost': 'oreCost',
        'Ruby Cost': 'rubyCost', 'Physical Def': 'physDef',
        'Elem. Block %': 'block',
    }, header_contains='Physical Def'),
    Table('Potions', 'potions', 'Region', 'region', {
        'Region': 'region', 'Rgn #': 'regionNum', 'Brew': 'brew',
        'Heals / Effect': 'heals', 'Ingredients': 'ingredients',
        'Ruby Cost': 'cost',
    }, section='HEALTH POTIONS'),
    Table('Potions', 'elixirs', 'Region', 'region', {
        'Region': 'region', 'Rgn #': 'regionNum', 'Brew': 'brew',
        'Heals / Effect': 'effect', 'Ingredients': 'ingredients',
        'Ruby Cost': 'cost', 'Immunity': 'immunity', 'Duration': 'duration',
    }, section='ELIXIRS'),
    Table('Guild', 'guildQuarry', 'Region', 'region', {
        'Rgn#': 'regionNum', 'N': 'N', 'Region': 'region',
        'Guild Quarry creature': 'creature', 'Base HP': 'baseHp',
        'Quarry HP (x3)': 'quarryHp', 'Head token': 'headToken',
    }, header_contains='Guild Quarry creature'),
]

# Prose sheets have no table to diff, so the only automatic check is: does the
# sentence that quotes a constant still quote the right number? Each entry pairs
# an export constant with the phrase it must appear inside — anchoring on the
# wording is what makes this worth running. A bare number search would pass on
# any coincidental match ("20" also occurs in "20x5 trophies"), which is false
# confidence, not a check.
#
# `{v}` is replaced with the value; the rest is a regex. A failure means either
# the constant moved or someone reworded the sentence — both worth a look.
# `scale` divides first, for constants the prose states in other units.
PROSE_CONSTANTS = {
    # Treasure & Chests is absent on purpose: its numbers are inline literals in
    # handlePickup, not named constants. See the design-workbook skill.
    'Quests': [
        ('collectorQty', r'{v} each of'),
        ('collectorTargets', r'each of {v} random'),
        ('chronicleStep', r'every {v} completed quests'),
    ],
    'NPCs & Services': [
        ('innRestCost', r'Full heal for {v} rubies'),
        ('portalToll', r'{v}-ruby toll'),
        ('villagerCount', r'~{v} wander'),
    ],
    'Potions': [('elixirImmunityMs', r'{v}s of FULL immunity', 1000)],
    'Guild': [('guildCullNeed', r'Slay {v} .* on a SINGLE hunt')],
    'Sword Progression': [('swordPartQty', r'prized trophies \({v} each\)')],
    'Armor Progression': [
        ('armorPartQty', r'signature trophies \({v} each\)'),
        ('armorSellValue', r'sells back for {v} rubies'),
    ],
}


# ─── Value comparison ─────────────────────────────────────────────────────────

def norm(value):
    """A comparable form: numbers as numbers, text stripped and unicode-folded.

    Excel round-trips whole floats as ints and vice versa, and the workbook mixes
    en/em dashes with hyphens, so neither type nor punctuation can be trusted.
    """
    if value is None:
        return ''
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else round(float(value), 6)
    text = unicodedata.normalize('NFKC', str(value)).strip()
    text = re.sub(r'\s+', ' ', text)
    text = text.replace('—', '-').replace('–', '-').replace('×', 'x')
    # A bare dash and an empty cell both mean "not applicable".
    return '' if text == '-' else text


def same(a, b):
    na, nb = norm(a), norm(b)
    if na == nb:
        return True
    # "35" from a formatted cell vs 35 from the code.
    try:
        return float(na) == float(nb)
    except (TypeError, ValueError):
        return False


def display(value):
    return '(blank)' if value is None or value == '' else repr(value)


# ─── Workbook walking ─────────────────────────────────────────────────────────

def find_header(ws, table):
    """Row index of the table's column-header row, or None."""
    want = table.header_contains
    start = 1
    if table.section:
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().startswith(table.section):
                start = r
                break
        else:
            return None
        # A banded section sits under the sheet's single column-header row, so
        # search backwards from the banner for it.
        for r in range(start, 0, -1):
            if any(norm(ws.cell(row=r, column=c).value) == norm(table.key)
                   for c in range(1, ws.max_column + 1)):
                return r
        return None
    for r in range(start, ws.max_row + 1):
        row = [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        if norm(want) in row and norm(table.key) in row:
            return r
    return None


def column_map(ws, header_row):
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            out[norm(v)] = c
    return out


def data_rows(ws, table, header_row, cols):
    """Sheet rows belonging to this table, as (row_index, key_value)."""
    key_col = cols[norm(table.key)]
    rows = []
    started = table.section is None
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(row=r, column=1).value
        label = norm(first)
        if table.section and not started:
            if isinstance(first, str) and first.strip().startswith(table.section):
                started = True
            continue
        if not started:
            continue
        # A banner (all-caps text spanning the row) ends the current table, as
        # does the notes block.
        other_cells = [ws.cell(row=r, column=c).value for c in range(2, ws.max_column + 1)]
        if isinstance(first, str) and not any(v is not None for v in other_cells):
            if rows:
                break
            continue
        key = ws.cell(row=r, column=key_col).value
        if key is None or norm(key) == '':
            if rows:
                break
            continue
        rows.append((r, key))
    return rows


# ─── The audit ────────────────────────────────────────────────────────────────

class Finding:
    def __init__(self, kind, sheet, text, cell=None, fix=None):
        self.kind, self.sheet, self.text, self.cell, self.fix = kind, sheet, text, cell, fix


def audit_table(ws, table, records, findings):
    header_row = find_header(ws, table)
    if header_row is None:
        findings.append(Finding('STRUCTURE', table.sheet,
                                f'could not find the header row for the "{table.key}" table'
                                + (f' under "{table.section}"' if table.section else '')
                                + ' — the sheet layout changed, update TABLES in this script'))
        return
    cols = column_map(ws, header_row)

    missing_cols = [h for h in table.columns if norm(h) not in cols]
    if missing_cols:
        findings.append(Finding('STRUCTURE', table.sheet,
                                f'columns not found in row {header_row}: {", ".join(missing_cols)}'))

    by_key = {}
    for rec in records:
        by_key.setdefault(norm(rec.get(table.key_field)), rec)

    seen = set()
    for row_idx, key_value in data_rows(ws, table, header_row, cols):
        k = norm(key_value)
        rec = by_key.get(k)
        if rec is None:
            findings.append(Finding('EXTRA', table.sheet,
                                    f'row {row_idx}: "{key_value}" is in the sheet but not in the code'))
            continue
        seen.add(k)
        for header, field in table.columns.items():
            if norm(header) not in cols:
                continue
            col = cols[norm(header)]
            cell = ws.cell(row=row_idx, column=col)
            expected = rec.get(field)
            if same(cell.value, expected):
                continue
            findings.append(Finding(
                'MISMATCH', table.sheet,
                f'{cell.coordinate} [{key_value} - {header}]: '
                f'sheet {display(cell.value)} but code says {display(expected)}',
                cell=cell,
                fix=expected if expected is not None else None))

    for k, rec in by_key.items():
        if k in seen:
            continue
        ordered = ' | '.join(
            f'{h}={rec.get(f) if rec.get(f) is not None else ""}'
            for h, f in table.columns.items())
        findings.append(Finding('MISSING', table.sheet,
                                f'"{rec.get(table.key_field)}" is in the code but has no row.'
                                f'\n      add: {ordered}'))


def audit_prose(ws, sheet, constants, checks, findings):
    text = ' '.join(
        str(c.value) for row in ws.iter_rows() for c in row if isinstance(c.value, str))
    text = re.sub(r'\s+', ' ', unicodedata.normalize('NFKC', text))
    for check in checks:
        key, template = check[0], check[1]
        scale = check[2] if len(check) > 2 else 1
        value = constants.get(key)
        if value is None:
            continue
        try:
            value = int(value) // scale if scale != 1 else value
        except (TypeError, ValueError):
            pass
        pattern = template.replace('{v}', re.escape(str(value)))
        if not re.search(pattern, text, re.IGNORECASE):
            findings.append(Finding(
                'CONSTANT', sheet,
                f'{key} is {value} in the source, but this sheet has no sentence matching '
                f'/{pattern}/ - either the prose still quotes the old number, or it was '
                f'reworded and this check needs updating'))


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--workbook', type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument('--fix', action='store_true',
                    help='rewrite mismatched cells in place (values only; styles are untouched)')
    ap.add_argument('--sheet', action='append', default=None,
                    help='limit to one sheet (repeatable)')
    ap.add_argument('--json', action='store_true', help='machine-readable findings')
    args = ap.parse_args()

    if not args.workbook.exists():
        sys.exit(f'design-audit: no workbook at {args.workbook}')

    try:
        raw = subprocess.run(['node', str(ROOT / 'tools' / 'design-export.js'), '--compact'],
                             capture_output=True, text=True, encoding='utf-8', check=True)
    except FileNotFoundError:
        sys.exit('design-audit: node is not on PATH')
    except subprocess.CalledProcessError as e:
        sys.exit('design-audit: the exporter failed -\n' + (e.stderr or '').strip())
    export = json.loads(raw.stdout)

    wb = openpyxl.load_workbook(args.workbook)
    findings = []

    for table in TABLES:
        if args.sheet and table.sheet not in args.sheet:
            continue
        if table.sheet not in wb.sheetnames:
            findings.append(Finding('STRUCTURE', table.sheet, 'sheet not found in the workbook'))
            continue
        audit_table(wb[table.sheet], table, export[table.source], findings)

    for sheet, checks in PROSE_CONSTANTS.items():
        if args.sheet and sheet not in args.sheet:
            continue
        if sheet in wb.sheetnames and checks:
            audit_prose(wb[sheet], sheet, export["constants"], checks, findings)

    fixed = 0
    if args.fix:
        for f in findings:
            if f.kind == 'MISMATCH' and f.cell is not None:
                f.cell.value = f.fix
                fixed += 1
        if fixed:
            wb.save(args.workbook)

    if args.json:
        print(json.dumps([{'kind': f.kind, 'sheet': f.sheet, 'detail': f.text}
                          for f in findings], indent=2))
    else:
        report(findings, args.workbook, fixed, bool(args.fix))

    unresolved = [f for f in findings if not (args.fix and f.kind == 'MISMATCH')]
    return 1 if unresolved else 0


def report(findings, workbook, fixed, fixing):
    order = ['STRUCTURE', 'MISSING', 'EXTRA', 'MISMATCH', 'CONSTANT']
    print(f'design-audit: {workbook}')
    if not findings:
        print('  clean - every derivable cell matches the game source.')
        return
    for kind in order:
        group = [f for f in findings if f.kind == kind]
        if not group:
            continue
        verb = 'fixed' if (fixing and kind == 'MISMATCH') else ''
        print(f'\n{kind} ({len(group)}){" - " + verb if verb else ""}')
        for f in group:
            print(f'  [{f.sheet}] {f.text}')
    print()
    if fixing:
        print(f'{fixed} cell(s) rewritten.')
    counts = {k: len([f for f in findings if f.kind == k]) for k in order}
    left = sum(v for k, v in counts.items() if not (fixing and k == 'MISMATCH'))
    if left:
        print(f'{left} finding(s) need a human: MISSING/EXTRA rows and prose are never '
              f'edited automatically.')


if __name__ == '__main__':
    sys.exit(main())
